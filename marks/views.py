from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib.auth.views import LoginView
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.contrib.auth import login
from django.urls import reverse
from datetime import datetime, timedelta, time
from django.db import transaction, IntegrityError
from django.db.models import Sum, Count, Q
from django.utils import timezone
from decimal import Decimal
from zoneinfo import ZoneInfo
import hashlib
import logging
import json
import re
import csv
import io
import secrets
from urllib.parse import urlencode
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


from .models import Bot, Branch, Tag, Product, PlanMonthly, Funnel, TrafficReport, PatchNote, UserProfile, TaskRequest, TikTokFunnelRequest, Experiment
from .forms import (
    BotForm,
    VKBotForm,
    BotStatusForm,
    BotDetailsForm,
    BranchForm,
    TagForm,
    CustomUserCreationForm,
    TagImportForm,
    PatchTaskRequestForm,
    MailingTaskRequestForm,
    BuildTaskRequestForm,
    TaskStatusForm,
    TikTokFunnelRequestForm,
)
from .experiment_forms import ExperimentCompletionForm, ExperimentForm
from .permissions import require_roles, BOT_OPERATORS_GROUP
from .services.telegram import (
    notify_new_task,
    notify_status_change,
    notify_done_to_user,
    notify_new_tiktok_funnel,
    send_weekly_tasks_report,
    send_text_message,
)
from .services.funnels import provision_funnel, activate_funnel
from .services.task_legacy import (
    get_task_feedback_map,
    get_task_tg_username,
    set_task_feedback_comment,
    set_task_tg_username,
)
from .task_time import build_task_day_window, format_task_datetime, get_tasks_timezone

logger = logging.getLogger(__name__)

TAG_UTM_FIELDS = ["utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"]
TAG_CLEAR_FIELDS = {field: None for field in TAG_UTM_FIELDS}
TAG_CLEAR_FIELDS["budget"] = None
TAG_CLEAR_FIELDS["url"] = None
TAG_UNDO_SESSION_KEY = "last_tag_action"
TELEGRAM_UPDATE_MESSAGE_KEYS = (
    "message",
    "edited_message",
    "channel_post",
    "edited_channel_post",
    "business_message",
    "edited_business_message",
)


def get_user_role(user):
    return getattr(getattr(user, "profile", None), "role", None)


def get_role_home_view_name(user):
    return "dashboard"


def _bot_sort_tuple(bot):
    created_at = getattr(bot, "created_at", None) or timezone.now()
    return (bot.platform_order, bot.sort_key, created_at)


def _branch_sort_tuple(branch):
    return (_bot_sort_tuple(branch.bot), (branch.name or "").casefold(), (branch.code or "").casefold())

def _active_tags_qs(tags_qs):
    return tags_qs.filter(url__isnull=False)

def _tag_snapshot(tag):
    snapshot = {field: getattr(tag, field) for field in TAG_UTM_FIELDS + ["budget", "url"]}
    if snapshot.get("budget") is not None:
        snapshot["budget"] = str(snapshot["budget"])
    return snapshot

def _set_last_tag_action(request, action, branch_id, payload):
    request.session[TAG_UNDO_SESSION_KEY] = {
        "action": action,
        "branch_id": branch_id,
        "payload": payload,
    }
    request.session.modified = True


def _flatten_form_errors(form):
    errors = []
    for field_errors in form.errors.values():
        errors.extend(str(error) for error in field_errors)
    return errors


def _serialize_tag_for_api(tag, branch=None, include_ab_key=False):
    payload = {
        "number": tag.number,
        "utm_source": tag.utm_source,
        "utm_medium": tag.utm_medium,
        "utm_campaign": tag.utm_campaign,
        "utm_term": tag.utm_term,
        "utm_content": tag.utm_content,
        "budget": tag.budget,
        "url": tag.url,
    }
    for field in TAG_UTM_FIELDS:
        if not payload.get(field):
            payload[field] = "None"

    if include_ab_key:
        payload["ab_key"] = _choose_branch_ab_key(branch or tag.branch)
    return payload


def _get_telegram_update_message(payload):
    for key in TELEGRAM_UPDATE_MESSAGE_KEYS:
        message = payload.get(key)
        if isinstance(message, dict):
            return message
    return {}


def _get_telegram_message_text(message):
    if not isinstance(message, dict):
        return ""
    return str(message.get("text") or message.get("caption") or "").strip()


def _extract_task_id_from_telegram_payload(value):
    if isinstance(value, str):
        match = re.search(r"#(\d+)\b", value)
        return int(match.group(1)) if match else None

    if isinstance(value, list):
        for item in value:
            task_id = _extract_task_id_from_telegram_payload(item)
            if task_id is not None:
                return task_id
        return None

    if not isinstance(value, dict):
        return None

    for key in ("text", "caption", "quote", "reply_to_message", "external_reply"):
        task_id = _extract_task_id_from_telegram_payload(value.get(key))
        if task_id is not None:
            return task_id

    for key, nested_value in value.items():
        if key in {"text", "caption", "quote", "reply_to_message", "external_reply"}:
            continue
        task_id = _extract_task_id_from_telegram_payload(nested_value)
        if task_id is not None:
            return task_id
    return None


def _extract_task_id_from_telegram_reply(message):
    if not isinstance(message, dict):
        return None
    for key in ("reply_to_message", "external_reply", "quote"):
        task_id = _extract_task_id_from_telegram_payload(message.get(key))
        if task_id is not None:
            return task_id
    return None


class RoleAwareLoginView(LoginView):
    template_name = "registration/login.html"


    def get_success_url(self):
        return reverse(get_role_home_view_name(self.request.user))


def user_is_admin(user):
    role = get_user_role(user)
    if user.is_superuser or role == UserProfile.Role.ADMIN:
        return True
    try:
        return user.groups.filter(name="Администратор").exists()
    except Exception:
        return False


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst')
def export_pdf(request):
    month = int(request.GET.get("month", datetime.now().month))
    year = int(request.GET.get("year", datetime.now().year))


    response = HttpResponse(content_type="application/pdf")
    filename = f"Отчёт_{month}_{year}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'


    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 100


    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, f"Отчёт по продуктам - {month}.{year}")
    y -= 30


    p.setFont("Helvetica", 11)
    for d in _get_dashboard_data(month, year):
        p.drawString(50, y, f"{d['product'].name}")
        y -= 20
        p.drawString(70, y, f"Расход: {d['spend']} руб. ({d['spend_delta']}%)")
        y -= 15
        p.drawString(70, y, f"Лиды: {d['leads']} ({d['leads_delta']}%)")
        y -= 25
        if y < 100:
            p.showPage()
            y = height - 100
            p.setFont("Helvetica", 11)
    p.save()
    return response


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst')
def export_excel(request):
    month = int(request.GET.get("month", datetime.now().month))
    year = int(request.GET.get("year", datetime.now().year))


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт {month}.{year}"
    ws.append(["Продукт", "Расход (руб.)", "Лиды", "Изм. расхода (%)", "Изм. лидов (%)"])


    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


    for d in _get_dashboard_data(month, year):
        ws.append([
            d["product"].name,
            d["spend"],
            d["leads"],
            f"{d['spend_delta']}%" if d["spend_delta"] else "-",
            f"{d['leads_delta']}%" if d["leads_delta"] else "-",
        ])


    for column_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Отчёт_{month}_{year}.xlsx"'
    wb.save(response)
    return response


def _get_dashboard_data(month, year):
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    products = Product.objects.all()
    def delta(curr, prev):
        if prev == 0:
            return None
        return round(((curr - prev) / prev) * 100, 1)
    data = []
    for product in products:
        reports = TrafficReport.objects.filter(product=product, month__month=month, month__year=year)
        prev_reports = TrafficReport.objects.filter(product=product, month__month=prev_month, month__year=prev_year)
        total_spend = reports.aggregate(Sum("spend"))["spend__sum"] or 0
        prev_spend = prev_reports.aggregate(Sum("spend"))["spend__sum"] or 0
        total_leads = (
            (reports.aggregate(Sum("leads_warm"))["leads_warm__sum"] or 0)
            + (reports.aggregate(Sum("leads_cold"))["leads_cold__sum"] or 0)
        )
        prev_leads = (
            (prev_reports.aggregate(Sum("leads_warm"))["leads_warm__sum"] or 0)
            + (prev_reports.aggregate(Sum("leads_cold"))["leads_cold__sum"] or 0)
        )
        data.append({
            "product": product,
            "spend": total_spend,
            "leads": total_leads,
            "spend_delta": delta(total_spend, prev_spend),
            "leads_delta": delta(total_leads, prev_leads),
        })
    return data


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def dashboard(request):
    return render(
        request,
        "marks/dashboard.html",
        {
            "is_admin_user": user_is_admin(request.user),
        },
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def update_field(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"error": "Некорректный JSON"}, status=400)


    model_type = data.get("model")
    record_id = data.get("id")
    field = data.get("field")
    value = data.get("value")


    model_map = {"plan": PlanMonthly, "report": TrafficReport, "tag": Tag}
    allowed_fields = {
        "plan": {"budget", "revenue_target", "warm_leads_target", "cold_leads_target", "notes"},
        "report": {"spend", "impressions", "clicks", "leads_warm", "leads_cold", "vendor", "notes", "platform", "month"},
        "tag": {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "budget"},
    }


    model = model_map.get(model_type)
    if not model:
        return JsonResponse({"error": "Недопустимая модель"}, status=400)
    if field not in allowed_fields.get(model_type, set()):
        return JsonResponse({"error": "Поле недоступно"}, status=400)


    try:
        obj = model.objects.get(id=record_id)
        previous_tag_snapshot = _tag_snapshot(obj) if model_type == "tag" else None
        model_field = obj._meta.get_field(field)
        itype = model_field.get_internal_type()


        def to_bool(v):
            if isinstance(v, bool):
                return v
            return str(v).lower() in {"1", "true", "yes", "on"}


        if itype in {"DecimalField"}:
            if model_type == "tag" and field == "budget" and (value is None or str(value).strip() == ""):
                coerced = None
            else:
                coerced = Decimal(value or 0)
        elif itype in {"IntegerField", "PositiveIntegerField", "BigIntegerField"}:
            coerced = int(value or 0)
        elif itype in {"DateField"}:
            s = (value or "").strip()
            if len(s) == 7:
                s = f"{s}-01"
            coerced = datetime.strptime(s, "%Y-%m-%d").date()
        elif itype in {"BooleanField"}:
            coerced = to_bool(value)
        else:
            coerced = value


        setattr(obj, field, coerced)
        obj.save(update_fields=[field])
        if model_type == "tag" and previous_tag_snapshot is not None:
            _set_last_tag_action(
                request,
                "edit",
                obj.branch_id,
                {"tag_id": obj.id, "fields": previous_tag_snapshot},
            )
        return JsonResponse({"success": True})
    except model.DoesNotExist:
        return JsonResponse({"error": "Объект не найден"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def duplicate_all_tags(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    count = int(request.POST.get("count", 1))
    tags = list(_active_tags_qs(branch.tags))
    total_created = 0
    created_ids = []


    for _ in range(count):
        for tag in tags:
            new_tag = Tag.objects.create(
                branch=branch,
                utm_source=tag.utm_source,
                utm_medium=tag.utm_medium,
                utm_campaign=tag.utm_campaign,
                utm_term=tag.utm_term,
                utm_content=tag.utm_content,
                budget=tag.budget,
            )
            total_created += 1
            created_ids.append(new_tag.id)
    if created_ids:
        _set_last_tag_action(
            request,
            "duplicate_all",
            branch.id,
            {"tag_ids": created_ids},
        )


    messages.success(request, f"Скопировано {total_created} меток ({len(tags)} x {count}).")
    return redirect("tags_list", branch_id=branch.id)


def register(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            profile = getattr(user, "profile", None)
            if profile is None:
                profile = UserProfile.objects.create(user=user)
            if profile.role != UserProfile.Role.BOT_USER:
                profile.role = UserProfile.Role.BOT_USER
                profile.save(update_fields=["role"])


            group, _ = Group.objects.get_or_create(name=BOT_OPERATORS_GROUP)
            user.groups.add(group)


            login(request, user)
            return redirect("bots_list")
    else:
        form = CustomUserCreationForm()
    return render(request, "registration/register.html", {"form": form})


def _get_active_branch_api_experiment(branch):
    today = timezone.localdate()
    experiments = (
        Experiment.objects.filter(
            branch=branch,
            wants_ab_test=True,
            status=Experiment.Status.IN_PROGRESS,
        )
        .order_by("-updated_at", "-id")
    )
    for experiment in experiments:
        if experiment.is_api_active(today):
            return experiment
    return None


def _choose_branch_ab_key(branch):
    experiment = _get_active_branch_api_experiment(branch)
    if experiment is None:
        return 1

    weight_a, weight_b = experiment.get_traffic_split_weights()
    total_weight = weight_a + weight_b
    return 1 if secrets.randbelow(total_weight) < weight_a else 2


def _build_branch_ab_payload(branch, assignment_key=""):
    experiment = _get_active_branch_api_experiment(branch)
    if experiment is None:
        return {"active": False}

    weight_a, weight_b = experiment.get_traffic_split_weights()
    total_weight = weight_a + weight_b
    assignment_key = (assignment_key or "").strip()
    if assignment_key:
        seed = f"{experiment.id}:{branch.id}:{assignment_key}".encode("utf-8")
        bucket = int(hashlib.sha256(seed).hexdigest()[:16], 16) % total_weight
        assignment_mode = "hash"
    else:
        bucket = secrets.randbelow(total_weight)
        assignment_mode = "random"

    variant_value = 1 if bucket < weight_a else 2
    return {
        "active": True,
        "experiment_id": experiment.id,
        "variant": "A" if variant_value == 1 else "B",
        "variant_value": variant_value,
        "split": f"{weight_a}/{weight_b}",
        "assignment_mode": assignment_mode,
    }


def bot_api(request, bot_name):
    normalized_name = (bot_name or "").strip()
    try:
        bot = Bot.objects.get(name=normalized_name)
    except Bot.DoesNotExist:
        if normalized_name.startswith("@"):
            try:
                telegram_bot = Bot.objects.get(
                    name=normalized_name.lstrip("@"),
                    platform=Bot.Platform.TELEGRAM,
                )
            except Bot.DoesNotExist:
                telegram_bot = None
            if telegram_bot is not None:
                return bot_api(request, normalized_name.lstrip("@"))
            return JsonResponse({"error": "Bot not found"}, status=404)
        return JsonResponse({"error": "Бот не найден"}, status=404)


    branch_code = (request.GET.get("branch_code") or "").strip()
    assignment_key = (request.GET.get("ab_key") or "").strip()

    filterable_fields = [
        "number",
        "utm_source",
        "utm_medium",
        "utm_campaign",
        "utm_term",
        "utm_content",
    ]
    utm_fields = ["utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"]
    tag_filters = {
        field: request.GET.get(field)
        for field in filterable_fields
        if request.GET.get(field)
    }
    number_only_request = bool(tag_filters) and set(tag_filters) == {"number"}


    data = {"bot": bot.name, "branches": []}
    filtered_tags = []
    branches_qs = bot.branches.all().prefetch_related("tags")
    if branch_code:
        branches_qs = branches_qs.filter(code__iexact=branch_code)
        if not branches_qs.exists():
            return JsonResponse({"error": "Branch not found"}, status=404)

    for branch in branches_qs:
        tags_qs = _active_tags_qs(branch.tags)
        if tag_filters:
            tags_qs = tags_qs.filter(**tag_filters)
        if number_only_request:
            tags_payload = [
                _serialize_tag_for_api(tag, branch=branch, include_ab_key=True)
                for tag in tags_qs
            ]
        else:
            tags_payload = list(
                tags_qs.values(
                    "number",
                    "utm_source",
                    "utm_medium",
                    "utm_campaign",
                    "utm_term",
                    "utm_content",
                    "budget",
                    "url",
                )
            )
            for tag in tags_payload:
                for field in utm_fields:
                    if not tag.get(field):
                        tag[field] = "None"


        if tag_filters:
            filtered_tags.extend(tags_payload)
            continue


        branch_data = {
            "name": branch.name,
            "code": branch.code,
            "tags": tags_payload,
        }
        if branch_code:
            branch_data["ab_test"] = _build_branch_ab_payload(branch, assignment_key=assignment_key)
        data["branches"].append(branch_data)


    if tag_filters:
        if not filtered_tags:
            return JsonResponse([], safe=False)
        if len(filtered_tags) == 1:
            return JsonResponse(filtered_tags[0])
        return JsonResponse(filtered_tags, safe=False)


    return JsonResponse(data, safe=False)


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def bots_list(request):
    bots = list(Bot.objects.all().annotate(branches_total=Count("branches")))
    bots.sort(key=_bot_sort_tuple)
    active_bots = [bot for bot in bots if bot.is_active]
    inactive_bots = [bot for bot in bots if not bot.is_active]
    if request.method == "POST":
        form_type = (request.POST.get("form_type") or "telegram").strip()
        telegram_form = BotForm(prefix="tg")
        vk_form = VKBotForm(prefix="vk")
        if form_type == "vk":
            vk_form = VKBotForm(request.POST, prefix="vk")
            if vk_form.is_valid():
                vk_form.save()
                messages.success(request, "VK бот создан.")
                return redirect("bots_list")
        else:
            telegram_form = BotForm(request.POST, prefix="tg")
            if telegram_form.is_valid():
                telegram_form.save()
                messages.success(request, "Telegram бот создан.")
                return redirect("bots_list")
    else:
        telegram_form = BotForm(prefix="tg")
        vk_form = VKBotForm(prefix="vk")
    return render(
        request,
        "marks/bots_list.html",
        {
            "active_bots": active_bots,
            "inactive_bots": inactive_bots,
            "telegram_form": telegram_form,
            "vk_form": vk_form,
        },
    )


@login_required
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def branches_list(request, bot_id):
    bot = get_object_or_404(Bot, id=bot_id)
    branches = sorted(
        bot.branches.all(),
        key=lambda branch: ((branch.name or "").casefold(), (branch.code or "").casefold()),
    )
    patchnotes = PatchNote.objects.filter(branch__bot=bot).select_related("branch")
    bot_status_form = BotStatusForm(bot=bot)
    bot_details_form = BotDetailsForm(instance=bot)
    form = BranchForm(bot=bot)
    if request.method == "POST":
        if request.POST.get("form_type") == "bot_status":
            bot_status_form = BotStatusForm(request.POST, bot=bot)
            if bot_status_form.is_valid():
                bot_status_form.save()
                messages.success(request, "Статус бота обновлён")
                return redirect("branches_list", bot_id=bot.id)
        elif request.POST.get("form_type") == "bot_details":
            bot_details_form = BotDetailsForm(request.POST, instance=bot)
            if bot_details_form.is_valid():
                bot_details_form.save()
                messages.success(request, "Информация о боте обновлена.")
                return redirect("branches_list", bot_id=bot.id)
        else:
            form = BranchForm(request.POST, bot=bot)
            if form.is_valid():
                branch = form.save(commit=False)
                branch.bot = bot
                branch.save()
                return redirect("branches_list", bot_id=bot.id)
    return render(
        request,
        "marks/branches_list.html",
        {
            "bot": bot,
            "branches": branches,
            "patchnotes": patchnotes,
            "form": form,
            "bot_status_form": bot_status_form,
            "bot_details_form": bot_details_form,
        },
    )


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def tags_list(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    tags = _active_tags_qs(branch.tags)
    patchnotes = branch.patch_notes.all()
    has_copied = bool(request.session.get("copied_tags"))


    if request.method == "POST" and "create_tag" in request.POST:
        if get_user_role(request.user) != 'analyst':
            form = TagForm(request.POST)
            if form.is_valid():
                tag = form.save(commit=False)
                tag.branch = branch
                tag.save()
                _set_last_tag_action(
                    request,
                    "create",
                    branch.id,
                    {"tag_ids": [tag.id]},
                )
                messages.success(request, "Метка создана")
                return redirect("tags_list", branch_id=branch.id)
    else:
        form = TagForm()


    import_form = TagImportForm()


    return render(
        request,
        "marks/tags_list.html",
        {
            "branch": branch,
            "tags": tags,
            "patchnotes": patchnotes,
            "form": form,
            "has_copied": has_copied,
            "import_form": import_form,
            "import_columns": TagImportForm.EXPECTED_COLUMNS,
        },
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def edit_tag(request, tag_id):
    tag = get_object_or_404(Tag, id=tag_id, url__isnull=False)
    previous_tag_snapshot = _tag_snapshot(tag)
    form = TagForm(request.POST, instance=tag)
    if form.is_valid():
        form.save()
        _set_last_tag_action(
            request,
            "edit",
            tag.branch_id,
            {"tag_id": tag.id, "fields": previous_tag_snapshot},
        )
        messages.success(request, f"Метка {tag.number} обновлена")
    else:
        messages.error(request, "Ошибка при обновлении метки")
    return redirect("tags_list", branch_id=tag.branch.id)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def copy_tags(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    copied = list(_active_tags_qs(branch.tags).values(
        "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "budget"
    ))
    for tag in copied:
        if tag.get("budget") is not None:
            tag["budget"] = str(tag["budget"])
    request.session["copied_tags"] = copied
    request.session.modified = True
    messages.success(request, "Таблица меток скопирована!")
    return redirect("tags_list", branch_id=branch.id)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def paste_tags(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    copied_tags = request.session.get("copied_tags")
    if not copied_tags:
        messages.error(request, "Нет скопированных меток.")
        return redirect("tags_list", branch_id=branch.id)
    created_ids = []
    for tag_data in copied_tags:
        new_tag = Tag.objects.create(branch=branch, **tag_data)
        created_ids.append(new_tag.id)
    if created_ids:
        _set_last_tag_action(
            request,
            "paste",
            branch.id,
            {"tag_ids": created_ids},
        )
    messages.success(request, "Таблица меток вставлена!")
    return redirect("tags_list", branch_id=branch.id)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def import_tags_csv(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    form = TagImportForm(request.POST, request.FILES)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect("tags_list", branch_id=branch.id)


    uploaded = form.cleaned_data["file"]
    uploaded.seek(0)
    expected = TagImportForm.EXPECTED_COLUMNS


    try:
        decoded = uploaded.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        messages.error(request, "Файл не прочитан в UTF-8.")
        return redirect("tags_list", branch_id=branch.id)


    reader = csv.DictReader(io.StringIO(decoded))
    headers = [(h or "").strip() for h in (reader.fieldnames or [])]
    if headers != expected:
        messages.error(
            request,
            "Структура CSV не подходит. Нужны столбцы: "
            + ", ".join(expected),
        )
        return redirect("tags_list", branch_id=branch.id)


    created = 0
    created_ids = []
    try:
        with transaction.atomic():
            for row in reader:
                if not any((row.get(col) or "").strip() for col in expected):
                    continue
                tag_kwargs = {col: (row.get(col) or "").strip() or None for col in expected}
                new_tag = Tag.objects.create(branch=branch, **tag_kwargs)
                created += 1
                created_ids.append(new_tag.id)
    except csv.Error as exc:
        messages.error(request, f"Ошибка CSV: {exc}")
        return redirect("tags_list", branch_id=branch.id)
    except Exception as exc:
        messages.error(request, f"Ошибка при импорте: {exc}")
        return redirect("tags_list", branch_id=branch.id)


    if created_ids:
        _set_last_tag_action(
            request,
            "import",
            branch.id,
            {"tag_ids": created_ids},
        )
    if created:
        messages.success(request, f"Метки добавлены: {created}.")
    else:
        messages.warning(request, "Подходящих строк в файле не нашлось.")
    return redirect("tags_list", branch_id=branch.id)


@login_required
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def duplicate_tag(request, tag_id):
    tag = get_object_or_404(Tag, id=tag_id, url__isnull=False)
    branch = tag.branch
    new_tag = Tag.objects.create(
        branch=branch,
        utm_source=tag.utm_source,
        utm_medium=tag.utm_medium,
        utm_campaign=tag.utm_campaign,
        utm_term=tag.utm_term,
        utm_content=tag.utm_content,
        budget=tag.budget,
    )
    _set_last_tag_action(
        request,
        "duplicate",
        branch.id,
        {"tag_ids": [new_tag.id]},
    )
    messages.success(request, f"Метка {new_tag.number} успешно создана как копия {tag.number}!")
    return redirect("tags_list", branch_id=branch.id)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def delete_tag(request, tag_id):
    tag = get_object_or_404(Tag, id=tag_id, url__isnull=False)
    previous_tag_snapshot = _tag_snapshot(tag)
    branch_id = tag.branch_id
    tag_number = tag.number
    Tag.objects.filter(id=tag.id).update(**TAG_CLEAR_FIELDS)
    _set_last_tag_action(
        request,
        "delete",
        branch_id,
        {"tag_id": tag.id, "fields": previous_tag_snapshot},
    )
    messages.success(request, f"Метка {tag_number} удалена.")
    return redirect("tags_list", branch_id=branch_id)


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', UserProfile.Role.BOT_USER)
def undo_tags_action(request, branch_id):
    action = request.session.get(TAG_UNDO_SESSION_KEY)
    if not action or action.get("branch_id") != branch_id:
        messages.error(request, "Нет изменений для отмены.")
        return redirect("tags_list", branch_id=branch_id)
    payload = action.get("payload") or {}
    action_type = action.get("action")
    if action_type in {"edit", "delete"}:
        tag_id = payload.get("tag_id")
        fields = payload.get("fields") or {}
        if tag_id and fields:
            Tag.objects.filter(id=tag_id, branch_id=branch_id).update(**fields)
            messages.success(request, "Последнее изменение отменено.")
        else:
            messages.error(request, "Нет данных для отмены изменения.")
    elif action_type in {"create", "duplicate", "duplicate_all", "paste", "import"}:
        tag_ids = payload.get("tag_ids") or []
        if tag_ids:
            Tag.objects.filter(id__in=tag_ids, branch_id=branch_id).update(**TAG_CLEAR_FIELDS)
            messages.success(request, "Последнее действие отменено.")
        else:
            messages.error(request, "Нет данных для отмены действия.")
    else:
        messages.error(request, "Нет изменений для отмены.")
    request.session.pop(TAG_UNDO_SESSION_KEY, None)
    request.session.modified = True
    return redirect("tags_list", branch_id=branch_id)


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst')
def product_reports(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    plans = PlanMonthly.objects.filter(product=product).order_by("-month")
    reports = TrafficReport.objects.filter(product=product).order_by("-month")


    if request.method == "POST" and get_user_role(request.user) != 'analyst':
        month = request.POST.get("month")
        platform = request.POST.get("platform")
        vendor = request.POST.get("vendor")
        spend = request.POST.get("spend")
        clicks = request.POST.get("clicks")
        leads_warm = request.POST.get("leads_warm")
        leads_cold = request.POST.get("leads_cold")


        TrafficReport.objects.create(
            product=product,
            month=month,
            platform=platform,
            vendor=vendor,
            spend=spend or 0,
            clicks=clicks or 0,
            leads_warm=leads_warm or 0,
            leads_cold=leads_cold or 0,
        )
        messages.success(request, "Отчёт успешно добавлен.")
        return redirect("product_reports", product_id=product.id)


    return render(request, "marks/product_reports.html", {"product": product, "plans": plans, "reports": reports})


def _parse_filter_date(value):
    value = (value or "").strip()
    if not value:
        return None, ""
    try:
        parsed = datetime.strptime(value, "%Y-%m-%d").date()
        return parsed, value
    except ValueError:
        return None, ""


def _get_task_filter_values(request):
    task_type = (request.GET.get("task_type") or "").strip()
    task_status = (request.GET.get("task_status") or "").strip()
    bot_id = (request.GET.get("bot_id") or "").strip()
    completed_from, completed_from_raw = _parse_filter_date(request.GET.get("completed_from"))
    completed_to, completed_to_raw = _parse_filter_date(request.GET.get("completed_to"))

    if task_type not in TaskRequest.Type.values:
        task_type = ""
    if task_status not in TaskRequest.Status.values:
        task_status = ""
    if not bot_id.isdigit():
        bot_id = ""

    if completed_from and completed_to and completed_from > completed_to:
        completed_from, completed_to = completed_to, completed_from
        completed_from_raw, completed_to_raw = completed_to_raw, completed_from_raw

    return {
        "task_type": task_type,
        "task_status": task_status,
        "bot_id": bot_id,
        "completed_from": completed_from,
        "completed_to": completed_to,
        "completed_from_raw": completed_from_raw,
        "completed_to_raw": completed_to_raw,
    }


def _apply_task_filters(tasks_qs, filters):
    if filters["task_type"]:
        tasks_qs = tasks_qs.filter(task_type=filters["task_type"])
    if filters["task_status"]:
        tasks_qs = tasks_qs.filter(status=filters["task_status"])
    if filters["bot_id"]:
        tasks_qs = tasks_qs.filter(branches__bot_id=int(filters["bot_id"]))
    dt_from, dt_to = build_task_day_window(filters["completed_from"], filters["completed_to"])
    if dt_from:
        tasks_qs = tasks_qs.filter(completed_at__gte=dt_from)
    if dt_to:
        tasks_qs = tasks_qs.filter(completed_at__lte=dt_to)
    return tasks_qs.distinct()


def _build_task_filter_query(filters):
    params = {}
    if filters.get("task_type"):
        params["task_type"] = filters["task_type"]
    if filters.get("task_status"):
        params["task_status"] = filters["task_status"]
    if filters.get("bot_id"):
        params["bot_id"] = filters["bot_id"]
    if filters.get("completed_from_raw"):
        params["completed_from"] = filters["completed_from_raw"]
    if filters.get("completed_to_raw"):
        params["completed_to"] = filters["completed_to_raw"]
    return urlencode(params)


def _to_report_dt_str(dt_value, report_tz):
    if not dt_value:
        return "-"
    try:
        return dt_value.astimezone(report_tz).strftime("%d.%m.%Y %H:%M")
    except Exception:
        return format_task_datetime(dt_value, empty="-")


def _build_completed_tasks_excel_bytes(tasks, report_tz):
    tasks = list(tasks)
    feedback_map = get_task_feedback_map(task.id for task in tasks)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выполненные задачи"
    ws.append([
        "ID",
        "Тип",
        "Статус",
        "Создал",
        "Создано",
        "Дедлайн",
        "Завершено",
        "Комментарий",
        "Фидбек",
        "CJM/ТЗ",
        "Бот и ветки",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for task in tasks:
        links = [value for value in [task.cjm_url, task.tz_url] if value]
        ws.append([
            task.id,
            task.get_task_type_display(),
            task.get_status_display(),
            task.created_by.username if task.created_by else "-",
            _to_report_dt_str(task.created_at, report_tz),
            _to_report_dt_str(task.deadline, report_tz),
            _to_report_dt_str(task.completed_at, report_tz),
            task.comment or "",
            feedback_map.get(task.id, ""),
            " | ".join(links),
            task.get_bot_branch_text(),
        ])

    for column_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 70)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), len(tasks)


def _week_to_date_window(base_date):
    days_since_sunday = (base_date.weekday() + 1) % 7
    sunday = base_date - timedelta(days=days_since_sunday)
    return sunday, base_date


def _month_to_date_window(base_date):
    return base_date.replace(day=1), base_date


def _handle_telegram_report_command(message_text, chat_id):
    if not message_text:
        return False

    parts = message_text.strip().split()
    if not parts:
        return False

    cmd = parts[0].split("@")[0].lower()
    if cmd not in {"/week", "/month"}:
        return False

    chat_id = str(chat_id or "").strip()
    if not chat_id:
        return True

    if len(parts) > 2:
        send_text_message(chat_id, "Формат: /week [YYYY-MM-DD] или /month [YYYY-MM-DD]")
        return True

    tz_name = (getattr(settings, "WEEKLY_TASKS_REPORT_TZ", "Europe/Moscow") or "Europe/Moscow").strip()
    try:
        report_tz = ZoneInfo(tz_name)
    except Exception:
        report_tz = ZoneInfo("UTC")

    if len(parts) == 2:
        try:
            base_date = datetime.strptime(parts[1], "%Y-%m-%d").date()
        except ValueError:
            send_text_message(chat_id, "Неверная дата. Используйте формат YYYY-MM-DD.")
            return True
    else:
        base_date = datetime.now(report_tz).date()

    if cmd == "/week":
        period_from, period_to = _week_to_date_window(base_date)
        caption_prefix = "текущую неделю"
        filename_prefix = "tasks_week_current"
    else:
        period_from, period_to = _month_to_date_window(base_date)
        caption_prefix = "текущий месяц"
        filename_prefix = "tasks_month_current"

    dt_from_local = datetime.combine(period_from, time.min, tzinfo=report_tz)
    dt_to_local_exclusive = datetime.combine(period_to + timedelta(days=1), time.min, tzinfo=report_tz)
    tasks_qs = (
        TaskRequest.objects.select_related("created_by")
        .prefetch_related("branches__bot")
        .filter(
            status=TaskRequest.Status.DONE,
            completed_at__isnull=False,
            completed_at__gte=dt_from_local,
            completed_at__lt=dt_to_local_exclusive,
        )
        .order_by("-completed_at")
    )

    content, tasks_count = _build_completed_tasks_excel_bytes(tasks_qs, report_tz)
    filename = f"{filename_prefix}_{period_from.isoformat()}_{period_to.isoformat()}.xlsx"
    caption = (
        f"Отчёт задачника за {caption_prefix} {period_from.strftime('%d.%m.%Y')} - "
        f"{period_to.strftime('%d.%m.%Y')}.\nВыполненных задач: {tasks_count}"
    )

    ok, error = send_weekly_tasks_report(
        chat_id=chat_id,
        filename=filename,
        content_bytes=content,
        caption=caption,
    )
    if not ok:
        send_text_message(chat_id, f"Не удалось отправить отчёт: {error}")
    return True


def _tasks_board_context(
    patch_form=None,
    mailing_form=None,
    build_form=None,
    tasks=None,
    filter_values=None,
    current_user=None,
):
    branch_options = sorted(Branch.objects.select_related("bot"), key=_branch_sort_tuple)
    branch_total = len(branch_options)
    patch_form = patch_form or PatchTaskRequestForm(prefix="patch")
    mailing_form = mailing_form or MailingTaskRequestForm(prefix="mailing")
    build_form = build_form or BuildTaskRequestForm(prefix="build")

    patch_selected = set(str(v) for v in (patch_form["branches"].value() or []))
    mailing_selected = set(str(v) for v in (mailing_form["branches"].value() or []))
    build_selected = set(str(v) for v in (build_form["branches"].value() or [])) if "branches" in build_form.fields else set()

    if tasks is None:
        tasks = list(
            TaskRequest.objects.select_related("created_by").prefetch_related("branches__bot").order_by("-created_at")
        )
    else:
        tasks = list(tasks)

    filter_values = filter_values or {
        "task_type": "",
        "task_status": "",
        "bot_id": "",
        "completed_from": None,
        "completed_to": None,
        "completed_from_raw": "",
        "completed_to_raw": "",
    }

    bot_filter_options = sorted(Bot.objects.filter(branches__isnull=False).distinct(), key=_bot_sort_tuple)
    completed_tasks_count = sum(1 for task in tasks if task.status == TaskRequest.Status.DONE and task.completed_at)
    done_tasks = [task for task in tasks if task.status == TaskRequest.Status.DONE and task.completed_at]
    completed_type_counters = {
        "build": sum(task.get_scope_units() for task in done_tasks if task.task_type == TaskRequest.Type.BUILD),
        "mailing": sum(task.get_scope_units() for task in done_tasks if task.task_type == TaskRequest.Type.MAILING),
        "patch": sum(task.get_scope_units() for task in done_tasks if task.task_type == TaskRequest.Type.PATCH),
    }

    columns = [
        {
            "status": TaskRequest.Status.UNREAD,
            "title": "Непрочитанное",
            "color": "secondary",
            "text": "text-white",
            "tasks": [],
        },
        {
            "status": TaskRequest.Status.IN_PROGRESS,
            "title": "В процессе",
            "color": "warning",
            "text": "text-dark",
            "tasks": [],
        },
        {
            "status": TaskRequest.Status.DONE,
            "title": "Готово",
            "color": "success",
            "text": "text-white",
            "tasks": [],
        },
    ]
    for task in tasks:
        for column in columns:
            if task.status == column["status"]:
                column["tasks"].append(task)
                break
    return {
        "patch_form": patch_form,
        "mailing_form": mailing_form,
        "build_form": build_form,
        "branch_options": branch_options,
        "patch_selected_branch_ids": patch_selected,
        "mailing_selected_branch_ids": mailing_selected,
        "build_selected_branch_ids": build_selected,
        "branch_total": branch_total,
        "task_type_choices": TaskRequest.Type.choices,
        "status_filter_choices": TaskRequest.Status.choices,
        "bot_filter_options": bot_filter_options,
        "filter_values": filter_values,
        "current_filter_query": _build_task_filter_query(filter_values),
        "completed_tasks_count": completed_tasks_count,
        "completed_type_counters": completed_type_counters,
        "kanban_columns": columns,
        "status_choices": TaskRequest.Status.choices,
        "show_kanban": user_is_admin(current_user) if current_user else False,
    }


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def tasks_board(request):
    filter_values = _get_task_filter_values(request)
    tasks_qs = TaskRequest.objects.select_related("created_by").prefetch_related("branches__bot").order_by("-created_at")
    filtered_tasks = _apply_task_filters(tasks_qs, filter_values)
    return render(
        request,
        "marks/tasks_board.html",
        _tasks_board_context(tasks=filtered_tasks, filter_values=filter_values, current_user=request.user),
    )


@login_required
@require_roles(UserProfile.Role.ADMIN)
def export_completed_tasks(request):
    filter_values = _get_task_filter_values(request)
    tasks_qs = TaskRequest.objects.select_related("created_by").prefetch_related("branches__bot").order_by("-completed_at")
    tasks_qs = _apply_task_filters(tasks_qs, filter_values).filter(
        status=TaskRequest.Status.DONE,
        completed_at__isnull=False,
    )
    tasks = list(tasks_qs)
    feedback_map = get_task_feedback_map(task.id for task in tasks)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выполненные задачи"
    ws.append([
        "ID",
        "Тип",
        "Статус",
        "Создал",
        "Создано",
        "Дедлайн",
        "Завершено",
        "Комментарий",
        "Фидбек",
        "CJM/ТЗ",
        "Бот и ветки",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for task in tasks:
        links = [value for value in [task.cjm_url, task.tz_url] if value]
        ws.append([
            task.id,
            task.get_task_type_display(),
            task.get_status_display(),
            task.created_by.username if task.created_by else "-",
            format_task_datetime(task.created_at),
            format_task_datetime(task.deadline),
            format_task_datetime(task.completed_at),
            task.comment or "",
            feedback_map.get(task.id, ""),
            " | ".join(links),
            task.get_bot_branch_text(),
        ])

    for column_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 70)

    period_parts = []
    if filter_values.get("completed_from_raw"):
        period_parts.append(f"from_{filter_values['completed_from_raw']}")
    if filter_values.get("completed_to_raw"):
        period_parts.append(f"to_{filter_values['completed_to_raw']}")
    suffix = "_".join(period_parts) if period_parts else "all"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="completed_tasks_{suffix}.xlsx"'
    wb.save(response)
    return response


def _handle_task_create(request, form, success_message, invalid_form_key):
    try:
        is_valid = form.is_valid()
    except Exception:
        logger.exception(
            "Task form validation crashed (form=%s, user_id=%s)",
            form.__class__.__name__,
            getattr(request.user, "id", None),
        )
        messages.error(
            request,
            "Произошла внутренняя ошибка при проверке формы заявки.",
        )
        return redirect("tasks_board")

    if not is_valid:
        logger.warning(
            "Task form invalid (form=%s, fields=%s, user_id=%s)",
            form.__class__.__name__,
            ",".join(sorted(form.errors.keys())),
            getattr(request.user, "id", None),
        )
        messages.error(request, "Не удалось создать задачу. Проверьте заполнение полей.")
        forms_state = {
            "patch_form": PatchTaskRequestForm(prefix="patch"),
            "mailing_form": MailingTaskRequestForm(prefix="mailing"),
            "build_form": BuildTaskRequestForm(prefix="build"),
        }
        forms_state[invalid_form_key] = form
        try:
            return render(
                request,
                "marks/tasks_board.html",
                _tasks_board_context(current_user=request.user, **forms_state),
            )
        except Exception:
            logger.exception("Failed to render tasks board with invalid form")
            return redirect("tasks_board")

    try:
        with transaction.atomic():
            task = form.save(commit=False)
            task.created_by = request.user
            task.status = TaskRequest.Status.UNREAD
            task.save()
            if hasattr(form, "save_m2m"):
                form.save_m2m()
            wants_notify = bool(form.cleaned_data.get("notify_me"))
            tg_username = form.cleaned_data.get("tg_username") or ""
            set_task_tg_username(task.id, tg_username if wants_notify else "")
    except Exception:
        logger.exception(
            "Failed to create task request (form=%s, user_id=%s)",
            form.__class__.__name__,
            getattr(request.user, "id", None),
        )
        messages.error(
            request,
            "Произошла внутренняя ошибка при создании заявки. Детали записаны в логах сервера.",
        )
        return redirect("tasks_board")

    notify_ok, notify_error = True, ""
    try:
        notify_result = notify_new_task(task)
        if isinstance(notify_result, tuple):
            notify_ok, notify_error = notify_result
        else:
            notify_ok, notify_error = bool(notify_result), ""
    except Exception:
        logger.exception("Failed to send new task notification (task_id=%s)", task.id)
        notify_ok, notify_error = False, "Сбой отправки уведомления."

    messages.success(request, success_message)
    if not notify_ok:
        messages.warning(
            request,
            "Задача создана, но уведомление в Telegram не отправлено. "
            + (notify_error or "Проверьте token/chat_id и перезапуск сервиса."),
        )
    return redirect("tasks_board")


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def create_patch_task(request):
    form = PatchTaskRequestForm(request.POST, request.FILES, prefix="patch")
    return _handle_task_create(
        request,
        form,
        "Заявка на правку создана.",
        "patch_form",
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def create_mailing_task(request):
    form = MailingTaskRequestForm(request.POST, request.FILES, prefix="mailing")
    return _handle_task_create(
        request,
        form,
        "Заявка на рассылку создана.",
        "mailing_form",
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def create_build_task(request):
    form = BuildTaskRequestForm(request.POST, request.FILES, prefix="build")
    return _handle_task_create(
        request,
        form,
        "Заявка на сборку бота создана.",
        "build_form",
    )


@login_required
@require_POST
@require_roles(UserProfile.Role.ADMIN)
def update_task_status(request, task_id):
    task = get_object_or_404(TaskRequest, id=task_id)
    old_status = task.status
    form = TaskStatusForm(request.POST, instance=task)
    if not form.is_valid():
        messages.error(request, "Не удалось обновить статус задачи.")
        return redirect("tasks_board")

    form.save()
    if old_status != task.status:
        notify_result = notify_status_change(task=task, old_status=old_status, changed_by=request.user)
        if isinstance(notify_result, tuple):
            notify_ok, notify_error = notify_result
        else:
            notify_ok, notify_error = bool(notify_result), ""
        user_notify_ok, user_notify_error = True, ""
        if task.status == TaskRequest.Status.DONE:
            tg_username = get_task_tg_username(task.id)
            if tg_username:
                user_notify_result = notify_done_to_user(task=task, tg_username=tg_username)
                if isinstance(user_notify_result, tuple):
                    user_notify_ok, user_notify_error = user_notify_result
                else:
                    user_notify_ok, user_notify_error = bool(user_notify_result), ""
        messages.success(request, "Статус задачи обновлён.")
        if not notify_ok:
            messages.warning(
                request,
                "Статус обновлён, но уведомление в Telegram не отправлено. "
                + (notify_error or "Проверьте token/chat_id и перезапуск сервиса."),
            )
        if not user_notify_ok:
            messages.warning(
                request,
                "Задача отмечена как выполненная, но персональное уведомление не отправлено. "
                + (user_notify_error or "Проверьте username/chat_id в заявке."),
            )
    else:
        messages.info(request, "Статус не изменился.")
    return redirect("tasks_board")


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def tiktok_funnels(request):
    funnels = list(TikTokFunnelRequest.objects.select_related("created_by"))
    return render(
        request,
        "marks/tiktok_funnels.html",
        {
            "form": TikTokFunnelRequestForm(),
            "funnels": funnels,
            "can_manage": user_is_admin(request.user),
        },
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def create_tiktok_funnel(request):
    form = TikTokFunnelRequestForm(request.POST)
    if not form.is_valid():
        funnels = list(TikTokFunnelRequest.objects.select_related("created_by"))
        messages.error(request, "Не удалось создать заявку. Проверьте поля.")
        return render(
            request,
            "marks/tiktok_funnels.html",
            {"form": form, "funnels": funnels, "can_manage": user_is_admin(request.user)},
        )

    try:
        funnel = form.save(commit=False)
        funnel.created_by = request.user
        funnel.status = TikTokFunnelRequest.Status.PENDING
        funnel.save()
    except IntegrityError:
        messages.error(request, "Воронка на этот endpoint уже существует.")
        return redirect("tiktok_funnels")
    except Exception:
        logger.exception("Failed to create TikTok funnel request (user_id=%s)", getattr(request.user, "id", None))
        messages.error(request, "Внутренняя ошибка при создании заявки. Детали в логах.")
        return redirect("tiktok_funnels")

    # Provision a pending draft row in the warehouse. A failure here must not
    # lose the local request — surface a warning and allow a later retry.
    ok, created, error = provision_funnel(
        landing_endpoint=funnel.landing_endpoint,
        offer=funnel.offer,
        bot_url=funnel.bot_url,
        bot_name=funnel.bot_name,
    )
    if ok:
        if funnel.warehouse_synced is False:
            TikTokFunnelRequest.objects.filter(pk=funnel.pk).update(warehouse_synced=True)
        if not created:
            messages.warning(
                request,
                "Заявка сохранена, но в складе уже была воронка на этот endpoint — черновик не перезаписан.",
            )
    else:
        messages.warning(
            request,
            "Заявка сохранена, но черновик в складе не записан. " + (error or "Повторите позже.")
            + " Можно повторить синхронизацию из списка.",
        )

    notify_ok, notify_error = True, ""
    try:
        notify_result = notify_new_tiktok_funnel(funnel)
        if isinstance(notify_result, tuple):
            notify_ok, notify_error = notify_result
    except Exception:
        logger.exception("Failed to send TikTok funnel notification (id=%s)", funnel.id)
        notify_ok, notify_error = False, "Сбой отправки уведомления."

    messages.success(request, "Заявка на TikTok-воронку создана.")
    if not notify_ok:
        messages.warning(request, "Уведомление в Telegram не отправлено. " + (notify_error or ""))
    return redirect("tiktok_funnels")


@login_required
@require_POST
@require_roles(UserProfile.Role.ADMIN)
def retry_tiktok_funnel_sync(request, funnel_id):
    funnel = get_object_or_404(TikTokFunnelRequest, id=funnel_id)
    ok, created, error = provision_funnel(
        landing_endpoint=funnel.landing_endpoint,
        offer=funnel.offer,
        bot_url=funnel.bot_url,
        bot_name=funnel.bot_name,
    )
    if ok:
        TikTokFunnelRequest.objects.filter(pk=funnel.pk).update(warehouse_synced=True)
        messages.success(request, "Черновик воронки записан в склад.")
    else:
        messages.error(request, "Не удалось записать в склад. " + (error or ""))
    return redirect("tiktok_funnels")


@login_required
@require_POST
@require_roles(UserProfile.Role.ADMIN)
def activate_tiktok_funnel(request, funnel_id):
    funnel = get_object_or_404(TikTokFunnelRequest, id=funnel_id)
    pixel_code = (request.POST.get("pixel_code") or "").strip()
    ok, error = activate_funnel(landing_endpoint=funnel.landing_endpoint, pixel_code=pixel_code)
    if ok:
        TikTokFunnelRequest.objects.filter(pk=funnel.pk).update(
            status=TikTokFunnelRequest.Status.ACTIVE,
            pixel_code=pixel_code,
        )
        messages.success(request, "Воронка активирована.")
    else:
        messages.error(request, "Активация не выполнена. " + (error or ""))
    return redirect("tiktok_funnels")


@csrf_exempt
@require_POST
def telegram_webhook(request, webhook_key):
    expected_key = (getattr(settings, "TELEGRAM_WEBHOOK_SECRET", "") or "").strip()
    if not expected_key or webhook_key != expected_key:
        return JsonResponse({"ok": False}, status=403)

    try:
        payload = json.loads((request.body or b"{}").decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "invalid_json"}, status=400)

    message = _get_telegram_update_message(payload)
    text = _get_telegram_message_text(message)
    chat = message.get("chat") or {}
    chat_id = chat.get("id")

    if text and _handle_telegram_report_command(text, chat_id):
        return JsonResponse({"ok": True})

    if not text:
        return JsonResponse({"ok": True})

    task_id = _extract_task_id_from_telegram_reply(message)
    if task_id is None:
        return JsonResponse({"ok": True})

    if not TaskRequest.objects.filter(id=task_id).exists():
        return JsonResponse({"ok": True})

    # Keep only the latest feedback text for a task.
    set_task_feedback_comment(task_id=task_id, feedback_comment=text[:4000])
    return JsonResponse({"ok": True})


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def _legacy_experiments_board(request):
    if request.method == "POST":
        form = ExperimentForm(request.POST)
        if form.is_valid():
            experiment = form.save(commit=False)
            experiment.created_by = request.user
            experiment.save()
            messages.success(request, "Эксперимент создан.")
            return redirect("experiments_board")
        messages.error(request, "Не удалось создать эксперимент. Проверьте заполнение полей.")
    else:
        form = ExperimentForm(initial={"status": Experiment.Status.BACKLOG})

    option_labels = dict(ExperimentForm.AB_TEST_OPTIONS)
    experiments = []
    for experiment in Experiment.objects.select_related("created_by", "branch__bot").all():
        selected_options = experiment.ab_test_options or []
        experiments.append(
            {
                "item": experiment,
                "ab_option_labels": [option_labels.get(code, code) for code in selected_options],
            }
        )

    return render(
        request,
        "marks/experiments_board.html",
        {
            "form": form,
            "experiments": experiments,
            "status_choices": Experiment.Status.choices,
        },
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def _legacy_update_experiment_status(request, experiment_id):
    experiment = get_object_or_404(Experiment, id=experiment_id)
    status_value = (request.POST.get("status") or "").strip()
    if status_value not in Experiment.Status.values:
        messages.error(request, "Недопустимый статус эксперимента.")
        return redirect("experiments_board")

    if experiment.status == status_value:
        messages.info(request, "Статус эксперимента не изменился.")
        return redirect("experiments_board")

    experiment.status = status_value
    experiment.save(update_fields=["status", "updated_at"])
    messages.success(request, "Статус эксперимента обновлен.")
    return redirect("experiments_board")


EXPERIMENT_ACTIVE_COLUMNS = [
    {"status": Experiment.Status.BACKLOG, "title": "Подготовка", "color": "secondary"},
    {"status": Experiment.Status.DRAFT, "title": "Разработка", "color": "primary"},
    {"status": Experiment.Status.IN_PROGRESS, "title": "В тесте", "color": "warning"},
    {"status": Experiment.Status.COMPLETED, "title": "На оценке", "color": "dark"},
]

EXPERIMENT_LIBRARY_COLUMNS = [
    {"status": Experiment.Status.SUCCESS, "title": "Успех", "color": "success"},
    {"status": Experiment.Status.FAILED, "title": "Провал", "color": "danger"},
    {"status": Experiment.Status.RETEST, "title": "Ретест", "color": "info"},
]

EXPERIMENT_FINAL_STATUSES = {column["status"] for column in EXPERIMENT_LIBRARY_COLUMNS}


EXPERIMENT_KIND_FILTER_CHOICES = [
    ("ab", "Только A/B"),
    ("plain", "Без A/B"),
]

DEFAULT_EXPERIMENT_DATALENS_URL = (
    "https://datalens.ru/8tid320t3bvmt-klienty-obshiy-dashbord"
    "?tab=Kx&state=1fe34562281&utm_referrer=about%3Ablank"
)


def _build_experiment_card(experiment, option_labels, default_dashboard_url):
    selected_options = experiment.ab_test_options or []
    return {
        "item": experiment,
        "ab_option_labels": [option_labels.get(code, code) for code in selected_options],
        "dashboard_href": (experiment.dashboard_url or "").strip() or default_dashboard_url,
    }


def _get_experiment_filter_values(request):
    status_value = (request.GET.get("status") or "").strip()
    bot_id = (request.GET.get("bot_id") or "").strip()
    branch_id = (request.GET.get("branch_id") or "").strip()
    kind = (request.GET.get("kind") or "").strip()

    return {
        "query": (request.GET.get("q") or "").strip(),
        "status": status_value if status_value in Experiment.Status.values else "",
        "bot_id": bot_id if bot_id.isdigit() else "",
        "branch_id": branch_id if branch_id.isdigit() else "",
        "kind": kind if kind in dict(EXPERIMENT_KIND_FILTER_CHOICES) else "",
    }


def _apply_experiment_filters(queryset, filter_values):
    query = (filter_values.get("query") or "").strip()
    if query:
        queryset = queryset.filter(
            Q(title__icontains=query)
            | Q(metric_impact__icontains=query)
            | Q(comparison_text__icontains=query)
            | Q(expected_change__icontains=query)
            | Q(hypothesis__icontains=query)
            | Q(comment__icontains=query)
            | Q(tz_url__icontains=query)
            | Q(branch__name__icontains=query)
            | Q(branch__code__icontains=query)
            | Q(branch__bot__name__icontains=query)
            | Q(branch__bot__display_name__icontains=query)
        )

    if filter_values.get("status"):
        queryset = queryset.filter(status=filter_values["status"])
    if filter_values.get("bot_id"):
        queryset = queryset.filter(branch__bot_id=filter_values["bot_id"])
    if filter_values.get("branch_id"):
        queryset = queryset.filter(branch_id=filter_values["branch_id"])

    kind = filter_values.get("kind")
    if kind == "ab":
        queryset = queryset.filter(wants_ab_test=True)
    elif kind == "plain":
        queryset = queryset.filter(wants_ab_test=False)

    return queryset


def _derive_experiment_task_deadline(experiment):
    for value in (experiment.start_date, experiment.duration_end_date, experiment.end_date):
        if value:
            return timezone.make_aware(datetime.combine(value, time(hour=12)), get_tasks_timezone())
    return timezone.now() + timedelta(days=3)


def _build_experiment_task_comment(experiment, option_labels):
    selected_options = experiment.ab_test_options or []
    parts = [f"Эксперимент: {experiment.title or f'#{experiment.id}'}"]

    if experiment.comparison_text:
        parts.append(f"Что с чем сравниваем: {experiment.comparison_text}")
    if selected_options:
        labels = [option_labels.get(code, code) for code in selected_options]
        parts.append(f"Что меняем: {', '.join(labels)}")
    if experiment.ab_test_custom_option:
        parts.append(f"Свой вариант: {experiment.ab_test_custom_option}")
    if experiment.metric_impact:
        parts.append(f"Цель: {experiment.metric_impact}")
    if experiment.expected_change:
        parts.append(f"Ожидаемое изменение: {experiment.expected_change}")
    if experiment.dashboard_url:
        parts.append(f"DataLens: {experiment.dashboard_url}")
    if experiment.comment:
        parts.append(f"Комментарий: {experiment.comment}")
    return " | ".join(parts)


def _unpack_notify_result(result):
    if isinstance(result, tuple):
        return bool(result[0]), result[1]
    return bool(result), ""


def _sync_experiment_technical_task(experiment, changed_by):
    option_labels = dict(ExperimentForm.AB_TEST_OPTIONS)
    task = experiment.technical_task
    is_created = task is None
    old_status = task.status if task is not None else ""

    task_payload = {
        "task_type": TaskRequest.Type.PATCH,
        "status": TaskRequest.Status.UNREAD,
        "cjm_url": "",
        "tz_url": (experiment.tz_url or "").strip(),
        "comment": _build_experiment_task_comment(experiment, option_labels),
        "deadline": _derive_experiment_task_deadline(experiment),
    }

    if is_created:
        task = TaskRequest.objects.create(created_by=changed_by, **task_payload)
    else:
        for field_name, field_value in task_payload.items():
            setattr(task, field_name, field_value)
        if task.created_by_id is None and changed_by is not None:
            task.created_by = changed_by
        task.save()

    if experiment.branch_id:
        task.branches.set([experiment.branch_id])
    else:
        task.branches.clear()

    if is_created or experiment.technical_task_id != task.id:
        experiment.technical_task = task
        experiment.save(update_fields=["technical_task", "updated_at"])

    if is_created:
        notify_result = notify_new_task(task)
    elif old_status != task.status:
        notify_result = notify_status_change(task=task, old_status=old_status, changed_by=changed_by)
    else:
        notify_result = notify_new_task(task)

    notify_ok, notify_error = _unpack_notify_result(notify_result)
    return task, is_created, notify_ok, notify_error


def _experiments_board_context(form, editing_experiment=None, filter_values=None):
    filter_values = filter_values or {
        "query": "",
        "status": "",
        "bot_id": "",
        "branch_id": "",
        "kind": "",
    }
    option_labels = dict(ExperimentForm.AB_TEST_OPTIONS)
    completion_form = ExperimentCompletionForm()
    default_dashboard_url = DEFAULT_EXPERIMENT_DATALENS_URL

    active_columns = [{**column, "items": []} for column in EXPERIMENT_ACTIVE_COLUMNS]
    library_columns = [{**column, "items": []} for column in EXPERIMENT_LIBRARY_COLUMNS]
    active_map = {column["status"]: column for column in active_columns}
    library_map = {column["status"]: column for column in library_columns}

    experiments_qs = Experiment.objects.select_related("created_by", "branch__bot", "technical_task").all()
    experiments_qs = _apply_experiment_filters(experiments_qs, filter_values)

    for experiment in experiments_qs:
        card = _build_experiment_card(
            experiment=experiment,
            option_labels=option_labels,
            default_dashboard_url=default_dashboard_url,
        )
        if experiment.status in active_map:
            active_map[experiment.status]["items"].append(card)
        elif experiment.status in library_map:
            library_map[experiment.status]["items"].append(card)

    return {
        "form": form,
        "editing_experiment": editing_experiment,
        "active_columns": active_columns,
        "library_columns": library_columns,
        "completion_form": completion_form,
        "default_dashboard_url": default_dashboard_url,
        "filter_values": filter_values,
        "status_filter_choices": Experiment.Status.choices,
        "kind_filter_choices": EXPERIMENT_KIND_FILTER_CHOICES,
        "bot_filter_options": sorted(Bot.objects.filter(branches__isnull=False).distinct(), key=_bot_sort_tuple),
        "branch_filter_options": sorted(Branch.objects.select_related("bot"), key=_branch_sort_tuple),
    }


@login_required
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def experiments_board(request):
    editing_experiment = None
    edit_id = (request.GET.get("edit") or "").strip()
    filter_values = _get_experiment_filter_values(request)

    if request.method == "POST":
        experiment_id = (request.POST.get("experiment_id") or "").strip()
        if experiment_id:
            editing_experiment = get_object_or_404(Experiment, id=experiment_id)
            form = ExperimentForm(request.POST, instance=editing_experiment)
        else:
            form = ExperimentForm(request.POST)

        if form.is_valid():
            experiment = form.save(commit=False)
            is_update = experiment.pk is not None
            if not is_update:
                experiment.created_by = request.user
            experiment.save()
            messages.success(
                request,
                "Эксперимент обновлен." if is_update else "Эксперимент создан.",
            )
            return redirect("experiments_board")

        messages.error(request, "Не удалось сохранить эксперимент. Проверьте заполнение полей.")
    else:
        if edit_id.isdigit():
            editing_experiment = get_object_or_404(Experiment, id=edit_id)
            form = ExperimentForm(instance=editing_experiment)
        else:
            form = ExperimentForm()

    return render(
        request,
        "marks/experiments_board.html",
        _experiments_board_context(
            form=form,
            editing_experiment=editing_experiment,
            filter_values=filter_values,
        ),
    )


@login_required
@require_POST
@require_roles('admin', 'manager', 'marketer', 'analyst', UserProfile.Role.BOT_USER)
def update_experiment_status(request, experiment_id):
    experiment = get_object_or_404(Experiment, id=experiment_id)
    status_value = (request.POST.get("status") or "").strip()
    completion_data = None
    if status_value not in Experiment.Status.values:
        messages.error(request, "Недопустимый статус эксперимента.")
        return redirect("experiments_board")

    if status_value == Experiment.Status.IN_PROGRESS and experiment.branch_id:
        if experiment.get_traffic_split_weights() is None:
            messages.error(request, "Для API A/B укажите сплит 50/50, 70/30 или свой в формате 80/20.")
            return redirect("experiments_board")

        conflict_exists = (
            Experiment.objects.filter(
                branch=experiment.branch,
                wants_ab_test=True,
                status=Experiment.Status.IN_PROGRESS,
            )
            .exclude(pk=experiment.pk)
            .exists()
        )
        if conflict_exists:
            messages.error(request, "Для этой ветки уже идет другой A/B тест.")
            return redirect("experiments_board")

    if status_value in EXPERIMENT_FINAL_STATUSES:
        completion_form = ExperimentCompletionForm(
            {
                "status": status_value,
                "start_date": request.POST.get("start_date") or experiment.start_date or "",
                "end_date": request.POST.get("end_date") or experiment.end_date or "",
                "dashboard_url": request.POST.get("dashboard_url")
                if "dashboard_url" in request.POST
                else experiment.dashboard_url,
                "result_variant_a": request.POST.get("result_variant_a")
                if "result_variant_a" in request.POST
                else experiment.result_variant_a,
                "result_variant_b": request.POST.get("result_variant_b")
                if "result_variant_b" in request.POST
                else experiment.result_variant_b,
                "comment": request.POST.get("comment") if "comment" in request.POST else experiment.comment,
            }
        )
        if not completion_form.is_valid():
            messages.error(
                request,
                "Перед финальным решением заполните: " + " ".join(_flatten_form_errors(completion_form)),
            )
            return redirect("experiments_board")
        completion_data = completion_form.cleaned_data
        status_value = completion_data["status"]

    if experiment.status == status_value:
        messages.info(request, "Статус эксперимента не изменился.")
        return redirect("experiments_board")

    task = None
    task_created = False
    notify_ok = True
    notify_error = ""
    with transaction.atomic():
        experiment.status = status_value
        update_fields = ["status", "updated_at"]
        if completion_data is not None:
            experiment.start_date = completion_data["start_date"]
            experiment.end_date = completion_data["end_date"]
            experiment.dashboard_url = completion_data["dashboard_url"]
            experiment.result_variant_a = completion_data["result_variant_a"]
            experiment.result_variant_b = completion_data["result_variant_b"]
            experiment.comment = completion_data["comment"]
            update_fields.extend(
                [
                    "start_date",
                    "end_date",
                    "dashboard_url",
                    "result_variant_a",
                    "result_variant_b",
                    "comment",
                ]
            )
        experiment.save(update_fields=update_fields)

        if status_value == Experiment.Status.DRAFT:
            task, task_created, notify_ok, notify_error = _sync_experiment_technical_task(
                experiment=experiment,
                changed_by=request.user,
            )

    messages.success(request, "Статус эксперимента обновлен.")
    if task is not None:
        if task_created:
            messages.success(request, f"Задача #{task.id} отправлена в техотдел.")
        else:
            messages.info(request, f"Задача #{task.id} в техотделе обновлена.")
        if not notify_ok:
            messages.warning(
                request,
                "Статус обновлен, но уведомление в техотдел не отправлено. "
                + (notify_error or "Проверьте Telegram-настройки."),
            )
    return redirect("experiments_board")
